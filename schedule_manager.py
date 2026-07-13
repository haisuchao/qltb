# schedule_manager.py
# Module quản lý lịch trực ban từ file Excel (Format mới Multi-sheet)

import pandas as pd
import os
import re
import glob
from datetime import datetime, timedelta
import config
from database import DatabaseManager


def get_schedule_filename(year):
    """Tên file chuẩn cho năm học bắt đầu từ 'year' (VD: year=2025 -> LichTrucBan_2025-2026.xlsx)"""
    return f"LichTrucBan_{year}-{year + 1}.xlsx"


SCHEDULE_FILENAME_PATTERN = re.compile(r'^LichTrucBan_(\d{4})-(\d{4})(?:_.*)?\.xlsx$')


class ScheduleManager:
    def __init__(self):
        self.db = DatabaseManager()
        self._seed_available_years_if_empty()

    def _seed_available_years_if_empty(self):
        """Nếu bảng available_years trống, quét thư mục lịch trực để đăng ký các file đã có sẵn
        (đảm bảo nâng cấp từ phiên bản cũ không làm mất cấu hình hiện có)."""
        if self.db.get_all_years():
            return

        found = []
        for filepath in glob.glob(os.path.join(config.SCHEDULE_FOLDER, "*.xlsx")):
            match = SCHEDULE_FILENAME_PATTERN.match(os.path.basename(filepath))
            if match:
                found.append((int(match.group(1)), os.path.basename(filepath)))

        if not found:
            return

        for year, filename in found:
            self.db.add_available_year(year, filename)

        latest_year = max(year for year, _ in found)
        self.db.set_current_year(latest_year)

    def get_schedule_sheet_name(self, date):
        """Lấy tên sheet tương ứng với tháng của ngày cần tra cứu"""
        # Format sheet name: m-yyyy (ví dụ: 8-2025)
        return f"{date.month}-{date.year}"

    def get_master_schedule_path(self):
        """Lấy đường dẫn file lịch trực tổng hợp (dựa vào năm học đang được quản lý trong DB)"""
        row = self.db.get_current_year_row()
        if row:
            year, filename = row
            path = os.path.join(config.SCHEDULE_FOLDER, filename)
            if os.path.exists(path):
                return path
            print(f"⚠️ File của năm hiện tại ({filename}) không tồn tại trên đĩa.")

        # Fallback cuối cùng: tìm file Excel bất kỳ trong thư mục (tránh 'chết cứng' nếu DB trống bất thường)
        print("⚠️ Không xác định được năm hiện tại từ DB, dùng fallback tìm file bất kỳ. Hãy dùng /set_current_year.")
        files = glob.glob(os.path.join(config.SCHEDULE_FOLDER, "*.xlsx"))
        if files:
            return files[0]
        return None
    
    def read_schedule_for_date(self, date):
        """Đọc lịch trực từ file Excel (Sheet tương ứng)"""
        filepath = self.get_master_schedule_path()
        if not filepath:
            print(f"Không tìm thấy file lịch trực trong {config.SCHEDULE_FOLDER}")
            return None
            
        sheet_name = self.get_schedule_sheet_name(date)
        
        try:
            # Kiểm tra xem sheet có tồn tại không
            xl = pd.ExcelFile(filepath)
            if sheet_name not in xl.sheet_names:
                # Thử format mm-yyyy nếu m-yyyy không có (ví dụ 08-2025)
                sheet_name_alt = f"{date.month:02d}-{date.year}"
                if sheet_name_alt in xl.sheet_names:
                    sheet_name = sheet_name_alt
                else:
                    print(f"Không tìm thấy sheet {sheet_name} trong file {filepath}")
                    return None
            
            # Đọc dữ liệu với header ở dòng 4 (index 3)
            # Columns: Ngày, Thứ, Trực ban 1 (Sáng), Trực ban 2 (Chiều), Trực lãnh đạo
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=3)
            
            # Chuẩn hóa tên cột để dễ xử lý (lấy theo index vì tên cột có thể thay đổi)
            # Index: 0=Date, 1=Day, 2=Morning, 3=Afternoon, 4=Leader
            if len(df.columns) < 5:
                print("File Excel không đủ số cột yêu cầu")
                return None
                
            # Đổi tên cột tạm thời để truy cập
            df.columns.values[0] = 'Date'
            df.columns.values[2] = 'Morning'
            df.columns.values[3] = 'Afternoon'
            df.columns.values[4] = 'Leader'
            
            # Xử lý ô gộp (Merged cells) cho cột Ngày: Điền giá trị từ trên xuống
            df['Date'] = df['Date'].ffill()
            
            print(f"📊 Đã đọc {len(df)} dòng dữ liệu từ sheet {sheet_name}")
            
            return df
        except Exception as e:
            print(f"Lỗi khi đọc file {filepath} (Sheet {sheet_name}): {str(e)}")
            return None
    
    def get_duty_info_for_date(self, date):
        """Lấy thông tin trực ban cho một ngày cụ thể"""
        df = self.read_schedule_for_date(date)
        
        if df is None:
            return None
        
        # Tìm dòng ứng với ngày
        # Chuyển đổi cột Date sang datetime để so sánh
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        
        target_date = pd.Timestamp(date.year, date.month, date.day)
        
        # Lọc dữ liệu
        duty_row = df[df['Date'] == target_date]
        
        if duty_row.empty:
            return None
        
        row = duty_row.iloc[0]
        
        # Xử lý các ô merged/empty (Weekend/Holiday)
        morning = row['Morning'] if pd.notna(row['Morning']) else None
        afternoon = row['Afternoon'] if pd.notna(row['Afternoon']) else None
        leader = row['Leader'] if pd.notna(row['Leader']) else None
        
        # Nếu không có ai trực (ngày nghỉ/lễ mà không phân công)
        if not morning and not afternoon and not leader:
             # Có thể trả về None hoặc object báo là ngày nghỉ
             return {
                'date': date.strftime('%d/%m/%Y'),
                'day_of_week': row.iloc[1] if pd.notna(row.iloc[1]) else '',
                'is_off': True,
                'morning_officer': None,
                'afternoon_officer': None,
                'leader': None
            }

        result = {
            'date': date.strftime('%d/%m/%Y'),
            'day_of_week': row.iloc[1] if pd.notna(row.iloc[1]) else '',
            'is_off': False,
            'morning_officer': str(morning).strip() if morning else None,
            'afternoon_officer': str(afternoon).strip() if afternoon else None,
            'leader': str(leader).strip() if leader else None
        }
        
        return result
    
    def get_tomorrow_duty(self):
        """Lấy thông tin trực ban ngày mai"""
        tomorrow = datetime.now() + timedelta(days=1)
        return self.get_duty_info_for_date(tomorrow)
    
    def update_schedule(self, date, shift, new_officer, old_officer=None, reason="", changed_by=""):
        """Cập nhật lịch trực (đổi người trực)"""
        if shift not in ['sáng', 'chiều']:
             print("Ca trực không hợp lệ")
             return False

        filepath = self.get_master_schedule_path()
        if not filepath:
            return False
            
        sheet_name = self.get_schedule_sheet_name(date)
        
        try:
             # Đọc file để lấy index dòng cần sửa
             # Lưu ý: openpyxl index bắt đầu từ 1.
             # Pandas read_excel header=3 nghĩa là dữ liệu bắt đầu từ row 5 (index 4 trong 0-based của list, nhưng row 5 trong Excel)
             
             # Cập nhật trực tiếp trên file local
             from openpyxl import load_workbook
             
             wb = load_workbook(filepath, data_only=True)
             if sheet_name not in wb.sheetnames:
                 # Check alt name
                 alt_name = f"{date.month:02d}-{date.year}"
                 if alt_name in wb.sheetnames:
                    sheet_name = alt_name
                 else:
                    return False
             
             ws = wb[sheet_name]
             
             # Tìm dòng chứa ngày
             target_row = None
             date_col_idx = 1 # Column A
             
             # Duyệt qua các dòng (bắt đầu từ dòng 5 do header là dòng 4)
             # So sánh dạng text để tránh lỗi định dạng
             target_date_str = date.strftime('%d/%m/%Y')
             
             for row in ws.iter_rows(min_row=5):
                 cell_val = row[0].value # Cột A
                 cell_date_str = None

                 if isinstance(cell_val, datetime):
                     cell_date_str = cell_val.strftime('%d/%m/%Y')
                 elif isinstance(cell_val, str):
                     # Chuẩn hóa về định dạng dd/mm/yyyy
                     cell_val_clean = cell_val.strip()
                     # Thử parse và format lại
                     for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%Y-%m-%d'):
                         try:
                             parsed = datetime.strptime(cell_val_clean, fmt)
                             cell_date_str = parsed.strftime('%d/%m/%Y')
                             break
                         except ValueError:
                             continue
                     # Nếu không parse được, so sánh trực tiếp
                     if not cell_date_str:
                         cell_date_str = cell_val_clean
                 print(f'cell_date_str: {cell_date_str}')
                 
                 if cell_date_str and cell_date_str == target_date_str:
                     target_row = row
                     break
            
             if not target_row:
                 print(f"Không tìm thấy ngày {date} trong sheet {sheet_name}")
                 return False
             
             # Xác định cột cần sửa
             # A=1, B=2, C=3 (Sáng), D=4 (Chiều)
             if shift == 'sáng':
                 cell_to_edit = target_row[2] # Cột C
                 current_val = cell_to_edit.value
             else:
                 cell_to_edit = target_row[3] # Cột D
                 current_val = cell_to_edit.value
             
             if old_officer is None:
                 old_officer = current_val
             
             # Cập nhật giá trị
             cell_to_edit.value = new_officer
             
             wb.save(filepath)
             
             # Log
             self.db.log_schedule_change(
                duty_date=date.strftime('%d/%m/%Y'),
                shift=shift,
                old_officer=str(old_officer) if old_officer else "N/A",
                new_officer=new_officer,
                reason=reason,
                approved_by=changed_by
            )
             
             return True
             
        except Exception as e:
            print(f"Lỗi update: {e}")
            return False

    def get_statistics(self, start_date, end_date):
        """Thống kê số buổi trực"""
        stats = {}
        # Duyệt qua các tháng trong khoảng thời gian
        # Để đơn giản, duyệt qua từng ngày (hơi chậm nếu range lớn nhưng chính xác)
        # Tối ưu: Duyệt qua từng tháng, đọc sheet tương ứng 1 lần
        
        current_month_start = start_date.replace(day=1)
        
        while current_month_start <= end_date:
            # Xác định sheet của tháng này
            df = self.read_schedule_for_date(current_month_start)
            
            if df is not None:
                # Lọc các ngày trong tháng nằm trong khoảng start-end
                df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
                mask = (df['Date'] >= pd.Timestamp(start_date)) & (df['Date'] <= pd.Timestamp(end_date))
                df_filtered = df[mask]
                
                for _, row in df_filtered.iterrows():
                     morning = row['Morning']
                     afternoon = row['Afternoon']
                     
                     if pd.notna(morning):
                         stats[morning] = stats.get(morning, 0) + 1
                     if pd.notna(afternoon):
                         stats[afternoon] = stats.get(afternoon, 0) + 1
            
            # Sang tháng tiếp theo
            if current_month_start.month == 12:
                current_month_start = current_month_start.replace(year=current_month_start.year+1, month=1)
            else:
                current_month_start = current_month_start.replace(month=current_month_start.month+1)
                
        return stats

    def generate_full_report(self):
        """Cập nhật bảng thống kê tổng hợp vào sheet 'Tổng' của file Excel
        (sheet này cũng chính là sheet được start_new_year khởi tạo ban đầu)"""
        filepath = self.get_master_schedule_path()
        if not filepath:
            return False, "Không tìm thấy file Excel"

        try:
            xl = pd.ExcelFile(filepath)
            # Lấy danh sách các sheet tháng (có format m-yyyy)
            month_sheets = [s for s in xl.sheet_names if '-' in s and s.split('-')[0].isdigit()]

            if not month_sheets:
                return False, "Không tìm thấy dữ liệu các tháng"

            # Sắp xếp các sheet theo thời gian (giả sử tên sheet là m-yyyy)
            def sheet_sort_key(s):
                m, y = map(int, s.split('-'))
                return y * 12 + m

            month_sheets.sort(key=sheet_sort_key)

            all_officers = set()
            monthly_data = {} # {sheet_name: {officer: count}}

            for sheet in month_sheets:
                # Đọc dữ liệu tháng
                df = pd.read_excel(filepath, sheet_name=sheet, header=3)
                if len(df.columns) < 4: continue

                counts = {}
                # Duyệt qua từng dòng để kiểm tra ô gộp (Nghỉ lễ/Tết)
                # Thông thường ô gộp sẽ có giá trị Sáng == Chiều
                for _, row in df.iterrows():
                    morning = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
                    afternoon = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""

                    # Nếu Sáng == Chiều và không rỗng -> Thường là ô gộp (Nghỉ lễ/Tết) -> Bỏ qua
                    if morning == afternoon and morning != "":
                        continue

                    # Danh sách các từ khóa cần bỏ qua (không phải tên người)
                    blacklist = ['x', '-', 'nghỉ', 'nan', 'thứ 7', 'chủ nhật', 'tết']

                    # Thống kê Sáng
                    if morning and not any(word in morning.lower() for word in blacklist):
                        counts[morning] = counts.get(morning, 0) + 1
                        all_officers.add(morning)

                    # Thống kê Chiều
                    if afternoon and not any(word in afternoon.lower() for word in blacklist):
                        counts[afternoon] = counts.get(afternoon, 0) + 1
                        all_officers.add(afternoon)

                monthly_data[sheet] = counts

            # Thứ tự cán bộ: theo "DS trực" (giữ nguyên STT); ai không có trong DS trực thì thêm cuối, không có STT
            roster = self._read_ds_truc_roster(filepath)
            roster_names = {o['name'] for o in roster}
            extra_officers = sorted(o for o in all_officers if o not in roster_names)
            ordered = [(o['stt'], o['name']) for o in roster] + [(None, name) for name in extra_officers]

            # Ghi vào sheet "Tổng": header dòng 5 (STT, Họ tên, các tháng dạng ngày, Tổng cộng), dòng cuối = tổng cột
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            wb = load_workbook(filepath)

            summary_sheet_name = "Tổng"
            if summary_sheet_name in wb.sheetnames:
                del wb[summary_sheet_name]

            # Tạo sheet mới ở vị trí đầu tiên
            ws = wb.create_sheet(summary_sheet_name, 0)

            ws.cell(row=5, column=1, value='STT').font = Font(bold=True)
            ws.cell(row=5, column=2, value='Họ tên').font = Font(bold=True)
            for j, sheet in enumerate(month_sheets, 3):
                m, y = map(int, sheet.split('-'))
                cell = ws.cell(row=5, column=j, value=datetime(y, m, 1))
                cell.number_format = 'mm/yyyy'
                cell.font = Font(bold=True)
            total_col = len(month_sheets) + 3
            ws.cell(row=5, column=total_col, value='Tổng cộng').font = Font(bold=True)

            row_idx = 6
            col_totals = [0] * len(month_sheets)
            for stt, name in ordered:
                ws.cell(row=row_idx, column=1, value=stt)
                ws.cell(row=row_idx, column=2, value=name)
                total = 0
                for j, sheet in enumerate(month_sheets):
                    count = monthly_data[sheet].get(name, 0)
                    ws.cell(row=row_idx, column=3 + j, value=count)
                    col_totals[j] += count
                    total += count
                ws.cell(row=row_idx, column=total_col, value=total)
                row_idx += 1

            # Dòng tổng cộng cuối bảng
            for j, col_total in enumerate(col_totals):
                ws.cell(row=row_idx, column=3 + j, value=col_total)
            ws.cell(row=row_idx, column=total_col, value=sum(col_totals))

            wb.save(filepath)
            return True, "Đã cập nhật bảng thống kê vào sheet 'Tổng'."

        except Exception as e:
            print(f"Lỗi generate report: {e}")
            return False, str(e)

    # def export_statistics_to_excel(self, start_date, end_date, output_file):

    def search_duty_schedule(self, name_query=None, date=None):
        """Tìm lịch trực trong tháng (theo ngày cung cấp hoặc tháng hiện tại).
        - name_query=None hoặc rỗng: Trả về toàn bộ lịch của tháng.
        - name_query có giá trị: Lọc theo tên.
        """
        if date is None:
            date = datetime.now()
            
        df = self.read_schedule_for_date(date)
        if df is None:
            return []
            
        results = []
        search_by_name = name_query and name_query.strip()
        if search_by_name:
            name_query = name_query.lower().strip()
        
        # Chuẩn hóa cột Date
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        
        for index, row in df.iterrows():
            duty_date = row['Date']
            
            # Debug: In giá trị ngày đọc được
            # print(f"DEBUG: Index {index}, Date: {duty_date}, Type: {type(duty_date)}")
            
            # Chuyển đổi sang datetime nếu chưa phải
            if not isinstance(duty_date, datetime) and pd.notna(duty_date):
                try:
                    # Thử parse chuỗi ngày (ưu tiên ngày trước tháng)
                    duty_date = pd.to_datetime(duty_date, dayfirst=True, errors='coerce')
                except:
                    pass

            # Nếu ko parse được ngày (NaT), bỏ qua dòng đó
            if pd.isna(duty_date):
                continue

            # Check Morning
            morning_raw = str(row.get('Morning', ''))
            morning = morning_raw.lower().strip()
            # Check Afternoon
            afternoon_raw = str(row.get('Afternoon', ''))
            afternoon = afternoon_raw.lower().strip()
            # Check Leader
            leader_raw = str(row.get('Leader', ''))
            leader = leader_raw.lower().strip()
            
            if search_by_name:
                # Lọc theo tên
                matched_role = []
                if name_query in morning:
                    matched_role.append('Sáng')
                if name_query in afternoon:
                    matched_role.append('Chiều')
                if name_query in leader:
                    matched_role.append('Lãnh đạo')
                    
                if matched_role:
                    res_item = {
                        'date': duty_date.strftime('%d/%m/%Y'),
                        'day_of_week': row.iloc[1] if pd.notna(row.iloc[1]) else '',
                        'roles': matched_role
                    }
                    print(f"✨ Tìm thấy: {res_item}")
                    results.append(res_item)
            else:
                # Khi xem toàn bộ lịch tháng:
                # Bỏ qua dòng nếu hoàn toàn không có ai trực
                blacklist = ['nan', '', 'x', '-']
                has_data = (
                    morning not in blacklist or
                    afternoon not in blacklist or
                    leader not in blacklist
                )
                if has_data:
                    res_item = {
                        'date': duty_date.strftime('%d/%m/%Y'),
                        'day_of_week': row.iloc[1] if pd.notna(row.iloc[1]) else '',
                        'morning': morning_raw.strip() if morning not in blacklist else '',
                        'afternoon': afternoon_raw.strip() if afternoon not in blacklist else '',
                        'leader': leader_raw.strip() if leader not in blacklist else '',
                    }
                    print(f"📅 Dòng dữ liệu: {res_item['date']} | S: {res_item['morning']} | C: {res_item['afternoon']} | LD: {res_item['leader']}")
                    results.append(res_item)
                    
        return results

    def swap_shifts(self, date1, shift1, date2, shift2, changed_by=""):
        """Đổi chỗ hai ca trực (có thể cùng ngày hoặc khác ngày)"""
        if shift1 not in ['sáng', 'chiều'] or shift2 not in ['sáng', 'chiều']:
             print("Ca trực không hợp lệ")
             return False, "Ca trực không hợp lệ (chỉ 'sáng' hoặc 'chiều')"

        filepath = self.get_master_schedule_path()
        if not filepath:
            return False, "Không tìm thấy file lịch trực"
            
        try:
            from openpyxl import load_workbook
            # Load with data_only=True to read formula values as dates
            wb = load_workbook(filepath, data_only=True)
            
            # Sheets for both dates
            sheet_name1 = self.get_schedule_sheet_name(date1)
            sheet_name2 = self.get_schedule_sheet_name(date2)
            
            if sheet_name1 not in wb.sheetnames or sheet_name2 not in wb.sheetnames:
                return False, "Không tìm thấy sheet tương ứng với tháng/năm"

            ws1 = wb[sheet_name1]
            ws2 = wb[sheet_name2]
            
            target_date_str1 = date1.strftime('%d/%m/%Y')
            target_date_str2 = date2.strftime('%d/%m/%Y')
            
            row1_idx = None
            row2_idx = None
            
            # Helper to normalize date and find row
            def find_row_idx(ws, target_str):
                for i, row in enumerate(ws.iter_rows(min_row=5), start=5):
                    cell_val = row[0].value
                    cell_date_str = None
                    if isinstance(cell_val, datetime):
                        cell_date_str = cell_val.strftime('%d/%m/%Y')
                    elif isinstance(cell_val, str):
                        cell_val_clean = cell_val.strip()
                        for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%Y-%m-%d'):
                            try:
                                cell_date_str = datetime.strptime(cell_val_clean, fmt).strftime('%d/%m/%Y')
                                break
                            except ValueError: continue
                        if not cell_date_str: cell_date_str = cell_val_clean
                    if cell_date_str == target_str:
                        return i
                return None

            row1_idx = find_row_idx(ws1, target_date_str1)
            row2_idx = find_row_idx(ws2, target_date_str2)

            if not row1_idx or not row2_idx:
                return False, "Không tìm thấy ngày trực trong lịch"

            # Identify columns: C=3 (morning), D=4 (afternoon)
            col1 = 3 if shift1 == 'sáng' else 4
            col2 = 3 if shift2 == 'sáng' else 4
            
            officer1 = ws1.cell(row=row1_idx, column=col1).value
            officer2 = ws2.cell(row=row2_idx, column=col2).value
            
            # Swap values
            ws1.cell(row=row1_idx, column=col1, value=officer2)
            ws2.cell(row=row2_idx, column=col2, value=officer1)
            
            wb.save(filepath)
            
            # Log changes
            self.db.log_schedule_change(
                duty_date=target_date_str1, shift=shift1,
                old_officer=str(officer1), new_officer=str(officer2),
                reason="Đổi chéo ca", approved_by=changed_by
            )
            self.db.log_schedule_change(
                duty_date=target_date_str2, shift=shift2,
                old_officer=str(officer2), new_officer=str(officer1),
                reason="Đổi chéo ca", approved_by=changed_by
            )
            
            return True, f"Đã đổi '{officer1}' ({target_date_str1} {shift1}) với '{officer2}' ({target_date_str2} {shift2})"
            
        except Exception as e:
            print(f"Lỗi swap: {e}")
            return False, str(e)

    def get_officer_list(self):
        """
        Đọc danh sách cán bộ từ sheet 'DS trực'
        Cấu trúc: Cột 2 (Họ tên), Cột 3 (Miễn trực - x)
        """
        filepath = self.get_master_schedule_path()
        if not filepath:
            return []
            
        try:
            # Đọc sheet 'DS trực', tiêu đề dòng 1
            df = pd.read_excel(filepath, sheet_name='DS trực', header=0)
            
            # Cột 2 là 'Họ tên', Cột 3 là 'Miễn trực' (index 1 và 2)
            # Lọc những người không có dấu 'x' ở cột Miễn trực
            valid_officers = []
            for _, row in df.iterrows():
                name = str(row.iloc[1]).strip()
                exemption = str(row.iloc[2]).strip().lower() if pd.notna(row.iloc[2]) else ""
                
                if name and name != "nan" and exemption != 'x':
                    valid_officers.append(name)
            
            return valid_officers
        except Exception as e:
            print(f"Lỗi đọc DS trực: {e}")
            return []

    def auto_generate_round_robin(self, month_year, names=None, leaders=None, start_name=None):
        """
        Tự động xếp lịch theo vòng tròn (Round-robin)
        - names: Nếu None, sẽ tự đọc từ sheet 'DS trực'
        - leaders: Danh sách lãnh đạo trực
        - start_name: Tên người bắt đầu xếp lịch (nếu có, sẽ bỏ qua logic nối tiếp tháng trước)
        - Bỏ qua Thứ 7, Chủ Nhật
        - Luân phiên sáng/chiều cho mỗi người
        """
        filepath = self.get_master_schedule_path()
        if not filepath:
            return False, "Không tìm thấy file Excel"

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Border, Side, Font
            import calendar
            
            # Phân tách month, year
            m, y = map(int, month_year.split('-'))
            last_day = calendar.monthrange(y, m)[1]
            
            wb = load_workbook(filepath)
            sheet_name = month_year
            
            if sheet_name in wb.sheetnames:
                # Nếu sheet đã tồn tại, xóa đi để tạo mới hoặc báo lỗi?
                # Ở đây ta sẽ ghi đè nội dung.
                ws = wb[sheet_name]
                # Xóa dữ liệu cũ từ dòng 5
                for row in ws.iter_rows(min_row=5):
                    for cell in row:
                        cell.value = None
            else:
                # Tạo sheet mới
                ws = wb.create_sheet(sheet_name)
            
            # Thiết lập header (Template)
            headers = ["Ngày", "Thứ", "Trực ban 1", "Trực ban 2", "Lãnh đạo trực"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Tiêu đề bảng (Dòng 1-2)
            ws.merge_cells('A1:E1')
            ws['A1'] = f"LỊCH TRỰC BAN THÁNG {m} NĂM {y}"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Logic xếp lịch
            idx_names = 0
            idx_leaders = 0
            use_start_name = False  # Cờ đánh dấu có dùng start_name hay không
            
            # Nếu names không được cung cấp, lấy từ DS trực
            if not names:
                names = self.get_officer_list()
                
            if not names:
                return False, "Không tìm thấy danh sách cán bộ trong sheet 'DS trực' hoặc danh sách trống."
            
            if not leaders:
                return False, "Danh sách lãnh đạo trực không được để trống."
            
            # Nếu có start_name, tìm vị trí trong danh sách và đặt idx_names
            if start_name:
                # Tìm kiếm tên (hỗ trợ tìm một phần tên, không phân biệt hoa thường)
                start_name_lower = start_name.lower().strip()
                found_idx = None
                for i, name in enumerate(names):
                    if start_name_lower in name.lower() or name.lower() in start_name_lower:
                        found_idx = i
                        break
                
                if found_idx is not None:
                    # Tính idx_names sao cho người đầu tiên trực sáng ngày đầu = start_name
                    # Với công thức m_idx = (idx + offset) % N, ta cần m_idx == found_idx khi idx_names == 0
                    # Đơn giản: idx_names = found_idx * (cần tính ngược từ công thức)
                    # Công thức sáng: m_idx = (idx_names + (idx_names // n_names if n_names % 2 == 0 else 0)) % n_names
                    # Khi idx_names nhỏ (< n_names), offset bổ sung = 0, nên m_idx = idx_names % n_names
                    # Vậy idx_names = found_idx
                    idx_names = found_idx
                    use_start_name = True
                    print(f"Bắt đầu xếp lịch từ: {names[found_idx]} (vị trí {found_idx})")
                else:
                    return False, f"Không tìm thấy '{start_name}' trong danh sách cán bộ."

            n_names = len(names)
            n_leaders = len(leaders)

            # --- LOGIC NỐI TIẾP THÁNG TRƯỚC (bỏ qua nếu đã chỉ định start_name) ---
            if not use_start_name:
                prev_m = m - 1 if m > 1 else 12
                prev_y = y if m > 1 else y - 1
                prev_sheet_name = f"{prev_m}-{prev_y}"
                
                if prev_sheet_name in wb.sheetnames:
                    ws_prev = wb[prev_sheet_name]
                    last_afternoon = None
                    
                    # Quét từ dưới lên để tìm người trực cuối cùng (Dòng 5 trở đi)
                    # max_row có thể chứa hàng trống, nên ta tìm hàng cuối có dữ liệu
                    for r in range(ws_prev.max_row, 4, -1):
                        val_a = ws_prev.cell(row=r, column=4).value # Afternoon
                        if val_a and not last_afternoon:
                            last_afternoon = str(val_a).strip()
                            break
                    
                    # Tìm vị trí của người cuối cùng trong danh sách hiện tại
                    if last_afternoon in names:
                        try:
                            last_idx = names.index(last_afternoon)
                            found_idx = 0
                            for test_idx in range(n_names * 2): # Quét đủ 1 vòng
                                 # Tính a_idx của test_idx
                                 off = (test_idx + 1) // n_names if n_names % 2 == 0 else 0
                                 if (test_idx + 1 + off) % n_names == last_idx:
                                     found_idx = test_idx + 2
                                     break
                            idx_names = found_idx
                        except: pass
                
            # -----------------------------------
            
            day_names = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
            
            row_idx = 5
            for day in range(1, last_day + 1):
                date_obj = datetime(y, m, day)
                weekday = date_obj.weekday() # 0=Monday, 6=Sunday
                
                # RESET LÃNH ĐẠO VÀO THỨ 2 HÀNG TUẦN
                if weekday == 0:
                    idx_leaders = 0
                
                # Cột A: Ngày (Ghi đối tượng datetime trực tiếp và set format)
                date_cell = ws.cell(row=row_idx, column=1, value=date_obj)
                date_cell.number_format = 'dd/mm/yyyy'
                
                # Cột B: Thứ
                ws.cell(row=row_idx, column=2, value=day_names[weekday])
                
                if weekday < 5: # Thứ 2 đến Thứ 6
                    # Công thức luân phiên: (idx + (idx // N if N%2==0 else 0)) % N
                    # Sáng
                    m_idx = (idx_names + (idx_names // n_names if n_names % 2 == 0 else 0)) % n_names
                    ws.cell(row=row_idx, column=3, value=names[m_idx])
                    
                    # Chiều
                    a_idx = (idx_names + 1 + ((idx_names + 1) // n_names if n_names % 2 == 0 else 0)) % n_names
                    ws.cell(row=row_idx, column=4, value=names[a_idx])
                    
                    # Lãnh đạo
                    ws.cell(row=row_idx, column=5, value=leaders[idx_leaders % n_leaders])
                    
                    idx_names += 2 # Tăng 2 slot (Sáng + Chiều)
                    idx_leaders += 1 # Tăng 1 slot cho Lãnh đạo
                else:
                    # Cuối tuần: Để trống hoặc ghi chú
                    ws.cell(row=row_idx, column=3, value="")
                    ws.cell(row=row_idx, column=4, value="")
                    ws.cell(row=row_idx, column=5, value="")
                
                row_idx += 1

            # Căn chỉnh và kẻ bảng
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
            for r in range(4, row_idx):
                for c in range(1, 6):
                    ws.cell(row=r, column=c).border = border
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 25

            wb.save(filepath)
            return True, f"Đã tự động xếp lịch xong cho tháng {month_year}."

        except Exception as e:
            print(f"Lỗi auto schedule: {e}")
            import traceback
            traceback.print_exc()
            return False, str(e)

    def _read_ds_truc_roster(self, filepath=None):
        """Đọc STT + Họ tên từ sheet 'DS trực' của file chỉ định (mặc định: file năm hiện tại)"""
        if filepath is None:
            filepath = self.get_master_schedule_path()
        if not filepath:
            return []

        try:
            df = pd.read_excel(filepath, sheet_name='DS trực', header=0)
            roster = []
            for _, row in df.iterrows():
                name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                if not name or name == "nan":
                    continue
                stt = row.iloc[0]
                if pd.notna(stt):
                    try:
                        stt = int(stt)
                    except (TypeError, ValueError):
                        pass
                else:
                    stt = None
                roster.append({'stt': stt, 'name': name})
            return roster
        except Exception as e:
            print(f"Lỗi đọc DS trực để copy sang năm mới: {e}")
            return []

    def start_new_year(self, year=None):
        """
        Tạo file Excel chuẩn cho năm học mới (tháng 8/year -> tháng 6/year+1).
        - Sheet 'DS trực': copy tên từ DS trực năm hiện tại, xóa cột Miễn/Lý do.
        - Sheet 'Tổng': liệt kê tên từ DS trực vừa tạo, số liệu = 0.
        - 11 sheet tháng: khung Ngày/Thứ, để trống Sáng/Chiều/Lãnh đạo (chờ /auto_schedule hoặc nhập tay).
        - Đăng ký năm mới vào available_years, cập nhật current_year = max(year, current hiện tại).
        Trả về (success, message, filepath)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Side, Font
        import calendar

        try:
            year = int(year) if year else datetime.now().year
        except (TypeError, ValueError):
            return False, f"Năm không hợp lệ: {year}", None

        if year < 2000 or year > 2100:
            return False, f"Năm không hợp lệ: {year}", None

        filename = get_schedule_filename(year)
        filepath = os.path.join(config.SCHEDULE_FOLDER, filename)

        if os.path.exists(filepath):
            return False, (
                f"⚠️ File '{filename}' đã tồn tại cho năm học {year}-{year + 1}. "
                f"Không tạo mới để tránh mất dữ liệu đã có. "
                f"Nếu muốn tạo lại từ đầu, hãy tự đổi tên/xóa file cũ trước, "
                f"hoặc dùng /set_current_year {year} nếu chỉ cần chuyển năm quản lý."
            ), None

        # Đọc danh sách cán bộ từ năm hiện tại TRƯỚC khi đổi current_year
        roster = self._read_ds_truc_roster()

        try:
            wb = Workbook()
            wb.remove(wb.active)  # Bỏ sheet mặc định "Sheet"

            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

            # --- Sheet "DS trực" ---
            ws_ds = wb.create_sheet('DS trực')
            for col, header in enumerate(['STT', 'Họ tên trực ban', 'Miễn', 'Lý do'], 1):
                cell = ws_ds.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

            for i, officer in enumerate(roster, 1):
                stt = officer['stt'] if officer['stt'] is not None else i
                ws_ds.cell(row=i + 1, column=1, value=stt)
                ws_ds.cell(row=i + 1, column=2, value=officer['name'])

            ws_ds.column_dimensions['A'].width = 8
            ws_ds.column_dimensions['B'].width = 30
            ws_ds.column_dimensions['C'].width = 10
            ws_ds.column_dimensions['D'].width = 25

            # Các tháng của năm học: 8/year -> 6/(year+1)
            months = [(8, year), (9, year), (10, year), (11, year), (12, year),
                      (1, year + 1), (2, year + 1), (3, year + 1), (4, year + 1),
                      (5, year + 1), (6, year + 1)]

            # --- Sheet "Tổng" (khớp layout thật: header dòng 5, cột tháng dạng ngày) ---
            ws_tong = wb.create_sheet('Tổng')
            ws_tong.cell(row=5, column=1, value='STT').font = Font(bold=True)
            ws_tong.cell(row=5, column=2, value='Họ tên').font = Font(bold=True)
            for j, (m, y) in enumerate(months, 3):
                cell = ws_tong.cell(row=5, column=j, value=datetime(y, m, 1))
                cell.number_format = 'mm/yyyy'
                cell.font = Font(bold=True)
            total_col = len(months) + 3
            ws_tong.cell(row=5, column=total_col, value='Tổng cộng').font = Font(bold=True)

            data_row = 6
            for i, officer in enumerate(roster, 1):
                stt = officer['stt'] if officer['stt'] is not None else i
                ws_tong.cell(row=data_row, column=1, value=stt)
                ws_tong.cell(row=data_row, column=2, value=officer['name'])
                for j in range(3, total_col + 1):
                    ws_tong.cell(row=data_row, column=j, value=0)
                data_row += 1

            # Dòng tổng cộng cuối bảng
            for j in range(3, total_col + 1):
                ws_tong.cell(row=data_row, column=j, value=0)

            # --- Các sheet tháng (khung trống, sẵn sàng cho /auto_schedule hoặc nhập tay) ---
            day_names = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
            for m, y in months:
                ws = wb.create_sheet(f"{m}-{y}")

                ws.merge_cells('A1:E1')
                ws['A1'] = f"LỊCH TRỰC BAN THÁNG {m} NĂM {y}"
                ws['A1'].font = Font(size=14, bold=True)
                ws['A1'].alignment = Alignment(horizontal='center')

                for col, header in enumerate(["Ngày", "Thứ", "Trực ban 1", "Trực ban 2", "Lãnh đạo trực"], 1):
                    cell = ws.cell(row=4, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                last_day = calendar.monthrange(y, m)[1]
                row_idx = 5
                for day in range(1, last_day + 1):
                    date_obj = datetime(y, m, day)
                    weekday = date_obj.weekday()

                    date_cell = ws.cell(row=row_idx, column=1, value=date_obj)
                    date_cell.number_format = 'dd/mm/yyyy'
                    ws.cell(row=row_idx, column=2, value=day_names[weekday])
                    ws.cell(row=row_idx, column=3, value="")
                    ws.cell(row=row_idx, column=4, value="")
                    ws.cell(row=row_idx, column=5, value="")
                    row_idx += 1

                for r in range(4, row_idx):
                    for c in range(1, 6):
                        ws.cell(row=r, column=c).border = border

                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 25

            os.makedirs(config.SCHEDULE_FOLDER, exist_ok=True)
            wb.save(filepath)

            # --- Đăng ký năm mới + cập nhật current_year ---
            existing_row = self.db.get_current_year_row()
            existing_current = existing_row[0] if existing_row else None

            self.db.add_available_year(year, filename)
            new_current = max(year, existing_current) if existing_current is not None else year
            self.db.set_current_year(new_current)

            return True, (
                f"Đã tạo file '{filename}' cho năm học {year}-{year + 1} "
                f"({len(roster)} cán bộ copy từ DS trực năm trước). "
                f"Năm hiện tại đang quản lý: {new_current}."
            ), filepath

        except Exception as e:
            print(f"Lỗi start_new_year: {e}")
            import traceback
            traceback.print_exc()
            return False, str(e), None
